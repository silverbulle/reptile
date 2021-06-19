# coding=gbk
import numpy as np
import xlrd
import xlwt
import openpyxl as op
import os
from tqdm import trange

f = open('data.txt', 'r', encoding='utf-8')  # 打开数据文本文档，注意编码格式的影响，这里用的是ANSI编码
w = open('data1.txt', 'w', encoding='utf-8')
bg = op.load_workbook(r"619.xlsx")
sheet = bg["Sheet1"]
# wb = xlwt.Workbook(encoding='uft-8')  # 打开一个excel文件
# ws1 = wb.add_sheet('first')  # 添加一个新表
row = 1  # 写入的起始行
col = 0  # 写入的起始列
k = 0
cnt = 1
data = f.read()
data = data.split('#')

# while '' in data:
#     data.remove('')
# print(data)
print(type(data), len(data))

# print(data[0], '\n', data[1])
# for i in trange(1, len(data)):
#     w.writelines(data[i-1])
#     if i % 8 == 0:
#         w.writelines('\n')

# f.close()
# w.close()


for row in trange(1, 47278):
    for column in range(1, 9):
        if cnt % 240 != 0:
            if data[k] == 'WY08':
                k += 1
            else:
                sheet.cell(row, column, data[k])
                k += 1
                cnt += 1
                # sheet.cell(row, column, 'silverbullet')
                # break
        else:
            # print('sdsaasfasfgsdafasdg')
            sheet.cell(row, column, 'None')
            cnt += 1
            break
        # print(k)
        # if k == len(data)-1:
            
bg.save("619.xlsx")
f.close()

# for lines in f:
#     a = lines.split('#')  # txt文件中每行的内容按‘ ’分割并存入数组中
#     k += 1
#     # rb = xlrd.open_workbook('C:\\Users\\DELL\\Desktop\\biao.xlsx')
#     # ws1 = rb.get_ws1(0)
#     for i in range(len(a)):
#         ws1.write(row, col, a[i])  # 向Excel文件中写入每一项
#         col += 1
#     row += 1
#     col = 0
# wb.save("619.xls")
